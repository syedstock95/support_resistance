import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import smtplib
from email.message import EmailMessage
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Config
EMAIL_FROM = "syedstock95@gmail.com"
EMAIL_TO = ["syedstock95@gmail.com", "2815699748@tmomail.net"]
EMAIL_SUBJECT = "?? Daily Market & Crypto News Summary"
EMAIL_APP_PASSWORD = "rkzh nbvt rhbw cztq"
OUTPUT_DIR = r"D:\\OneDrive\\Documents\\shares\\Document\\News"
TODAY = datetime.now().strftime("%Y-%m-%d")
FMP_API = "cn2AZpCgLd44PYFPCHVkmqZouGukDFXL"

# News sources
NEWS_SOURCES = {
    "Yahoo Finance RSS": "https://finance.yahoo.com/news/rssindex",
    "Google News - Business": "https://news.google.com/rss/search?q=business&hl=en-US&gl=US&ceid=US:en",
    "CoinDesk": "https://www.coindesk.com/arc/outboundfeeds/rss/",
    "Cointelegraph": "https://cointelegraph.com/rss",
    "CNBC": "https://www.cnbc.com/id/100003114/device/rss/rss.html"
}

FMP_ENDPOINTS = [
    (f"https://financialmodelingprep.com/api/v4/general_news?page=0&apikey={FMP_API}", "FMP General News"),
    (f"https://financialmodelingprep.com/api/v4/stock-news-sentiments-rss-feed?page=0&apikey={FMP_API}", "FMP Stock Sentiment"),
    (f"https://financialmodelingprep.com/api/v4/crypto_news?page=0&apikey={FMP_API}", "FMP Crypto News")
]

# Fetch RSS
def fetch_rss_news(name, url):
    headlines = []
    try:
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(resp.content, 'xml')
        for item in soup.find_all('item')[:10]:
            headlines.append({
                'source': name,
                'title': item.title.text.strip(),
                'link': item.link.text.strip(),
                'summary': item.description.text[:300] if item.description else '',
                'date': TODAY
            })
    except Exception as e:
        print(f"Error fetching {name}: {e}")
    return headlines

# Fetch FMP API
def fetch_fmp_api_news(url, name):
    headlines = []
    try:
        resp = requests.get(url)
        data = resp.json()
        for item in data:
            pub_date = item.get("publishedDate", item.get("date", ""))
            if pub_date.startswith(TODAY):
                headlines.append({
                    'source': name,
                    'title': item.get("title", "").strip(),
                    'link': item.get("url", "").strip(),
                    'summary': item.get("text", "")[:300],
                    'date': pub_date
                })
    except Exception as e:
        print(f"Error fetching {name}: {e}")
    return headlines

# Compile news
def compile_news():
    all_news = []
    for name, url in NEWS_SOURCES.items():
        all_news.extend(fetch_rss_news(name, url))
    for url, name in FMP_ENDPOINTS:
        all_news.extend(fetch_fmp_api_news(url, name))
    return all_news

# Save news to Excel with formatting
def save_to_excel(news_items):
    df = pd.DataFrame(news_items)
    now = datetime.now().strftime("%Y%m%d_%H%M")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    file_path = os.path.join(OUTPUT_DIR, f"market_news_summary_{now}.xlsx")
    df.to_excel(file_path, index=False)

    # Apply formatting
    wb = load_workbook(file_path)
    ws = wb.active
    blue_bold_font = Font(bold=True, color="0000FF")  # Blue bold font

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if cell.row != 1:  # skip header row
                    cell.font = blue_bold_font
        ws.column_dimensions[col_letter].width = max_length + 5

    for cell in ws[1]:  # Header row
        cell.font = Font(bold=True, color="000000")  # Bold black header

    wb.save(file_path)
    print(f"? Excel saved to: {file_path}")
    return file_path

# Email sender
def send_email(news_items, attachment=None):
    msg = EmailMessage()
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"] = EMAIL_FROM
    msg["To"] = ', '.join(EMAIL_TO)

    content = "\n".join([f"- {item['title']} ({item['source']})\n{item['link']}" for item in news_items[:5]])
    msg.set_content(f"?? Top Headlines ({TODAY}):\n\n{content}")

    if attachment:
        with open(attachment, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application', subtype='octet-stream', filename=os.path.basename(attachment))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_FROM, EMAIL_APP_PASSWORD)
            smtp.send_message(msg)
        print("?? Email and SMS sent!")
    except Exception as e:
        print(f"? Failed to send email: {e}")

# Run everything
if __name__ == "__main__":
    news_items = compile_news()
    if news_items:
        file = save_to_excel(news_items)
        send_email(news_items, file)
    else:
        print("?? No news found for today.")
